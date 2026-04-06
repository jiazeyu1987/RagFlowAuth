import os
import unittest

from authx import TokenPayload
from fastapi import FastAPI, Request
from fastapi.testclient import TestClient

from backend.app.core import auth as auth_module
from backend.app.modules.preview.router import router as preview_router
from backend.tests._util_tempdir import cleanup_dir, make_temp_dir


class _User:
    def __init__(self, role: str = "viewer", group_ids: list[int] | None = None):
        self.user_id = "u1"
        self.username = "u1"
        self.full_name = "测试用户"
        self.email = "u1@example.com"
        self.role = role
        self.status = "active"
        self.company_id = 1
        self.group_id = (group_ids or [None])[0]
        self.group_ids = list(group_ids or [])


class _UserStore:
    def __init__(self, user: _User):
        self._user = user

    def get_by_user_id(self, user_id: str):  # noqa: ARG002
        return self._user


class _KbDoc:
    def __init__(self, doc_id: str, kb_id: str, file_path: str, filename: str):
        self.doc_id = doc_id
        self.kb_id = kb_id
        self.file_path = file_path
        self.filename = filename


class _KbStore:
    def __init__(self, doc: _KbDoc):
        self._doc = doc

    def get_document(self, doc_id: str):
        if doc_id == self._doc.doc_id:
            return self._doc
        return None


class _RagflowService:
    def __init__(self, content: bytes, filename: str):
        self._content = content
        self._filename = filename

    def download_document(self, doc_id: str, dataset: str):  # noqa: ARG002
        return self._content, self._filename


class _PermissionGroupStore:
    def get_group(self, group_id: int):
        if group_id == 1:
            return {
                "can_upload": False,
                "can_review": False,
                "can_download": True,
                "can_copy": False,
                "can_delete": False,
                "can_manage_kb_directory": False,
                "can_view_kb_config": False,
                "can_view_tools": False,
                "accessible_kbs": ["kb1"],
                "accessible_chats": [],
                "accessible_tools": [],
            }
        return None


class _Company:
    def __init__(self, name: str = "测试公司"):
        self.name = name


class _OrgDirectoryStore:
    def get_company(self, company_id: int):
        if company_id == 1:
            return _Company()
        return None


class _WatermarkPolicy:
    policy_id = "wm-default"
    name = "默认水印策略"
    text_template = "用户:{username} | 公司:{company} | 时间:{timestamp} | 用途:{purpose} | 文档ID:{doc_id}"
    label_text = "受控预览"
    text_color = "#6b7280"
    opacity = 0.18
    rotation_deg = -24
    gap_x = 260
    gap_y = 180
    font_size = 18


class _WatermarkPolicyStore:
    def get_active_policy(self):
        return _WatermarkPolicy()


class _Deps:
    def __init__(self, kb_doc: _KbDoc):
        self.user_store = _UserStore(_User(role="viewer", group_ids=[1]))
        self.kb_store = _KbStore(kb_doc)
        self.ragflow_service = _RagflowService(b"%PDF-1.4 test", "x.pdf")
        self.permission_group_store = _PermissionGroupStore()
        self.org_directory_store = _OrgDirectoryStore()
        self.org_structure_manager = self.org_directory_store
        self.watermark_policy_store = _WatermarkPolicyStore()
        self.knowledge_directory_manager = None


def _override_get_current_payload(_: Request) -> TokenPayload:
    return TokenPayload(sub="u1")


def _write_simple_xlsx(path: str) -> None:
    import openpyxl
    import openpyxl.worksheet._writer as _writer

    orig_remove = _writer.os.remove

    def safe_remove(p: str) -> None:
        try:
            orig_remove(p)
        except PermissionError:
            return

    _writer.os.remove = safe_remove  # type: ignore[assignment]
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "c1"
        ws["B1"] = "c2"
        ws["A2"] = 1
        ws["B2"] = 2
        wb.save(path)
    finally:
        _writer.os.remove = orig_remove  # type: ignore[assignment]


class TestPreviewGatewayUnit(unittest.TestCase):
    def test_gateway_knowledge_returns_text_contract(self):
        td = make_temp_dir(prefix="ragflowauth_preview_gateway")
        try:
            path = os.path.join(str(td), "a.txt")
            with open(path, "wb") as f:
                f.write("hello".encode("utf-8"))

            kb_doc = _KbDoc(doc_id="k1", kb_id="kb1", file_path=path, filename="a.txt")

            app = FastAPI()
            app.state.deps = _Deps(kb_doc)
            app.include_router(preview_router, prefix="/api")
            app.dependency_overrides[auth_module.get_current_payload] = _override_get_current_payload

            with TestClient(app) as client:
                resp = client.get("/api/preview/documents/knowledge/k1/preview")

            self.assertEqual(resp.status_code, 200)
            data = resp.json()
            self.assertEqual(data.get("type"), "text")
            self.assertEqual(data.get("filename"), "a.txt")
            self.assertIn("content", data)
            self.assertEqual(data.get("watermark", {}).get("policy_id"), "wm-default")
            self.assertIn("用途:预览", data.get("watermark", {}).get("text", ""))
        finally:
            cleanup_dir(td)

    def test_gateway_knowledge_markdown_returns_text_contract(self):
        td = make_temp_dir(prefix="ragflowauth_preview_gateway")
        try:
            path = os.path.join(str(td), "a.md")
            with open(path, "wb") as f:
                f.write("# title\n\n- a\n".encode("utf-8"))

            kb_doc = _KbDoc(doc_id="k_md", kb_id="kb1", file_path=path, filename="a.md")

            app = FastAPI()
            app.state.deps = _Deps(kb_doc)
            app.include_router(preview_router, prefix="/api")
            app.dependency_overrides[auth_module.get_current_payload] = _override_get_current_payload

            with TestClient(app) as client:
                resp = client.get("/api/preview/documents/knowledge/k_md/preview")

            self.assertEqual(resp.status_code, 200, resp.text)
            data = resp.json()
            self.assertEqual(data.get("type"), "text")
            self.assertEqual(data.get("filename"), "a.md")
            self.assertIn("# title", data.get("content", ""))
        finally:
            cleanup_dir(td)

    def test_gateway_knowledge_csv_returns_text_contract(self):
        td = make_temp_dir(prefix="ragflowauth_preview_gateway")
        try:
            path = os.path.join(str(td), "a.csv")
            with open(path, "wb") as f:
                f.write("c1,c2\n1,2\n".encode("utf-8"))

            kb_doc = _KbDoc(doc_id="k_csv", kb_id="kb1", file_path=path, filename="a.csv")

            app = FastAPI()
            app.state.deps = _Deps(kb_doc)
            app.include_router(preview_router, prefix="/api")
            app.dependency_overrides[auth_module.get_current_payload] = _override_get_current_payload

            with TestClient(app) as client:
                resp = client.get("/api/preview/documents/knowledge/k_csv/preview")

            self.assertEqual(resp.status_code, 200, resp.text)
            data = resp.json()
            self.assertEqual(data.get("type"), "text")
            self.assertEqual(data.get("filename"), "a.csv")
            self.assertIn("c1,c2", data.get("content", ""))
        finally:
            cleanup_dir(td)

    def test_gateway_ragflow_returns_pdf_contract(self):
        app = FastAPI()
        kb_doc = _KbDoc(doc_id="k1", kb_id="kb1", file_path=__file__, filename="a.txt")
        app.state.deps = _Deps(kb_doc)
        app.include_router(preview_router, prefix="/api")
        app.dependency_overrides[auth_module.get_current_payload] = _override_get_current_payload

        with TestClient(app) as client:
            resp = client.get("/api/preview/documents/ragflow/r1/preview?dataset=kb1")

        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data.get("type"), "pdf")
        self.assertEqual(data.get("filename"), "x.pdf")
        self.assertIn("content", data)
        self.assertIn("文档ID:r1", data.get("watermark", {}).get("text", ""))

    def test_gateway_ragflow_requires_dataset(self):
        app = FastAPI()
        kb_doc = _KbDoc(doc_id="k1", kb_id="kb1", file_path=__file__, filename="a.txt")
        app.state.deps = _Deps(kb_doc)
        app.include_router(preview_router, prefix="/api")
        app.dependency_overrides[auth_module.get_current_payload] = _override_get_current_payload

        with TestClient(app) as client:
            resp = client.get("/api/preview/documents/ragflow/r1/preview")

        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.json()["detail"], "missing_dataset")

    def test_gateway_knowledge_excel_default_is_table_mode(self):
        td = make_temp_dir(prefix="ragflowauth_preview_gateway")
        try:
            path = os.path.join(str(td), "a.xlsx")
            _write_simple_xlsx(path)

            kb_doc = _KbDoc(doc_id="k2", kb_id="kb1", file_path=path, filename="a.xlsx")

            app = FastAPI()
            app.state.deps = _Deps(kb_doc)
            app.include_router(preview_router, prefix="/api")
            app.dependency_overrides[auth_module.get_current_payload] = _override_get_current_payload

            with TestClient(app) as client:
                resp = client.get("/api/preview/documents/knowledge/k2/preview")

            self.assertEqual(resp.status_code, 200, resp.text)
            data = resp.json()
            self.assertEqual(data.get("type"), "excel")
            self.assertEqual(data.get("filename"), "a.xlsx")
            self.assertIn("sheets", data)
        finally:
            cleanup_dir(td)

    def test_gateway_knowledge_excel_render_html_is_html(self):
        td = make_temp_dir(prefix="ragflowauth_preview_gateway")
        try:
            path = os.path.join(str(td), "a.xlsx")
            _write_simple_xlsx(path)

            kb_doc = _KbDoc(doc_id="k3", kb_id="kb1", file_path=path, filename="a.xlsx")

            app = FastAPI()
            app.state.deps = _Deps(kb_doc)
            app.include_router(preview_router, prefix="/api")
            app.dependency_overrides[auth_module.get_current_payload] = _override_get_current_payload

            with TestClient(app) as client:
                resp = client.get("/api/preview/documents/knowledge/k3/preview?render=html")

            self.assertEqual(resp.status_code, 200, resp.text)
            data = resp.json()
            self.assertEqual(data.get("type"), "html")
            self.assertTrue(str(data.get("filename", "")).endswith(".html"))
            self.assertIn("content", data)
        finally:
            cleanup_dir(td)
