from __future__ import annotations

import base64
import io
import os
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from fastapi import UploadFile

from backend.app.core.managed_paths import resolve_managed_child_storage_path
from backend.app.core.paths import resolve_repo_path
from .store import PackageDrawingStore


_MODEL_ALIASES = {"型号", "model", "产品型号", "modelno", "modelnumber", "规格型号"}
_BARCODE_ALIASES = {"条形码", "条码", "barcode", "bar_code", "ean", "code"}
_IMAGE_ALIASES = {"示意图", "图片", "图", "image", "img", "picture", "photo"}

_IMAGE_EXT_BY_MIME = {
    "image/png": "png",
    "image/jpeg": "jpg",
    "image/gif": "gif",
    "image/bmp": "bmp",
    "image/webp": "webp",
}

_IMAGE_MIME_BY_FORMAT = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "gif": "image/gif",
    "bmp": "image/bmp",
    "webp": "image/webp",
}


@dataclass(frozen=True)
class _RowImage:
    source_type: str
    image_url: str
    rel_path: str
    mime_type: str
    filename: str
    content: bytes


class PackageDrawingImportError(Exception):
    def __init__(self, code: str):
        super().__init__(code)
        self.code = code


class PackageDrawingManager:
    def __init__(self, deps: Any):
        self._store: PackageDrawingStore = deps.package_drawing_store
        self._image_root = resolve_repo_path("data/package_drawing_images")
        self._image_root.mkdir(parents=True, exist_ok=True)

    async def import_from_upload(self, upload_file: UploadFile) -> dict[str, Any]:
        filename = str(getattr(upload_file, "filename", "") or "").strip()
        if not filename:
            raise PackageDrawingImportError("file_required")
        if not filename.lower().endswith(".xlsx"):
            raise PackageDrawingImportError("only_xlsx_supported")

        content = await upload_file.read()
        if not content:
            raise PackageDrawingImportError("empty_file")

        parse_result = self._parse_workbook(content)
        errors: list[dict[str, Any]] = list(parse_result["errors"])
        success = 0

        for model, payload in parse_result["records"].items():
            row_images = payload["images"]
            saved_abs_paths: list[Path] = []
            db_images: list[dict[str, Any]] = []
            order_index = 0
            try:
                for item in row_images:
                    if item.source_type == "url":
                        db_images.append(
                            {
                                "image_id": uuid.uuid4().hex,
                                "source_type": "url",
                                "image_url": item.image_url,
                                "rel_path": "",
                                "mime_type": "",
                                "filename": "",
                                "sort_order": order_index,
                            }
                        )
                        order_index += 1
                        continue

                    image_id = uuid.uuid4().hex
                    ext = _IMAGE_EXT_BY_MIME.get(item.mime_type, "bin")
                    model_dir = self._safe_model_dir_name(model)
                    rel_path = str(Path(model_dir) / f"{image_id}.{ext}")
                    abs_path = self._image_root / rel_path
                    abs_path.parent.mkdir(parents=True, exist_ok=True)
                    abs_path.write_bytes(item.content)
                    saved_abs_paths.append(abs_path)

                    db_images.append(
                        {
                            "image_id": image_id,
                            "source_type": "embedded",
                            "image_url": "",
                            "rel_path": rel_path.replace("\\", "/"),
                            "mime_type": item.mime_type,
                            "filename": item.filename,
                            "sort_order": order_index,
                        }
                    )
                    order_index += 1

                old_paths = self._store.upsert_record(
                    model=model,
                    barcode=payload["barcode"],
                    parameters=payload["parameters"],
                    images=db_images,
                )
                self._cleanup_old_paths(old_paths)
                success += 1
            except Exception as exc:
                for p in saved_abs_paths:
                    try:
                        if p.exists():
                            p.unlink()
                    except Exception:
                        pass
                errors.append(
                    {
                        "sheet": payload["sheet"],
                        "row": payload["row"],
                        "model": model,
                        "reason": f"save_failed:{exc}",
                    }
                )

        total = len(parse_result["records"])
        failed = len(errors)
        return {
            "filename": filename,
            "rows_scanned": parse_result["rows_scanned"],
            "total": total,
            "success": success,
            "failed": failed,
            "errors": errors,
        }

    def query_by_model(self, model: str) -> dict[str, Any] | None:
        clean_model = str(model or "").strip()
        if not clean_model:
            return None
        record = self._store.get_record(clean_model)
        if record is None:
            return None

        images: list[dict[str, Any]] = []
        for item in record.images:
            if item.source_type == "url" and item.image_url:
                images.append(
                    {
                        "type": "url",
                        "url": item.image_url,
                    }
                )
                continue

            if item.source_type != "embedded" or not item.rel_path:
                continue
            abs_path = resolve_managed_child_storage_path(
                item.rel_path,
                managed_root=self._image_root,
                field_name="package_drawing_images.rel_path",
            )
            if not abs_path.exists():
                continue
            raw = abs_path.read_bytes()
            data_url = f"data:{item.mime_type or 'application/octet-stream'};base64,{base64.b64encode(raw).decode('ascii')}"
            images.append(
                {
                    "type": "embedded",
                    "data_url": data_url,
                    "mime_type": item.mime_type or "application/octet-stream",
                    "filename": item.filename or abs_path.name,
                    "image_id": item.image_id,
                }
            )

        return {
            "model": record.model,
            "barcode": record.barcode or "",
            "parameters": record.parameters or {},
            "images": images,
        }

    def get_image_binary(self, image_id: str) -> tuple[bytes, str, str] | None:
        clean_id = str(image_id or "").strip()
        if not clean_id:
            return None
        image = self._store.get_image(clean_id)
        if image is None:
            return None
        if image.source_type != "embedded" or not image.rel_path:
            return None
        abs_path = resolve_managed_child_storage_path(
            image.rel_path,
            managed_root=self._image_root,
            field_name="package_drawing_images.rel_path",
        )
        if not abs_path.exists():
            return None
        filename = image.filename or abs_path.name
        mime_type = image.mime_type or "application/octet-stream"
        return abs_path.read_bytes(), filename, mime_type

    def _parse_workbook(self, content: bytes) -> dict[str, Any]:
        try:
            import openpyxl
        except Exception as exc:
            raise PackageDrawingImportError(f"excel_engine_unavailable:{exc}") from exc

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise PackageDrawingImportError(f"invalid_xlsx:{exc}") from exc

        records_by_model: dict[str, dict[str, Any]] = {}
        errors: list[dict[str, Any]] = []
        rows_scanned = 0

        for ws in wb.worksheets:
            header_values = [self._text(v) for v in self._iter_header(ws)]
            max_columns = max(int(ws.max_column or 0), len(header_values))
            if max_columns <= 0:
                continue

            header_names = {
                idx: self._normalize_header_name(header_values[idx - 1] if idx - 1 < len(header_values) else "", idx)
                for idx in range(1, max_columns + 1)
            }
            model_col = self._find_fixed_column(header_names, _MODEL_ALIASES)
            barcode_col = self._find_fixed_column(header_names, _BARCODE_ALIASES)
            image_col = self._find_fixed_column(header_names, _IMAGE_ALIASES)
            if model_col is None:
                errors.append(
                    {
                        "sheet": ws.title,
                        "row": 1,
                        "model": "",
                        "reason": "missing_model_column",
                    }
                )
                continue

            embedded_by_row = self._extract_embedded_images_by_row(ws)
            for row_idx in range(2, int(ws.max_row or 0) + 1):
                row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_columns + 1)]
                if self._row_is_empty(row_values):
                    continue
                rows_scanned += 1

                model = self._text(row_values[model_col - 1] if model_col - 1 < len(row_values) else "")
                if not model:
                    errors.append(
                        {
                            "sheet": ws.title,
                            "row": row_idx,
                            "model": "",
                            "reason": "missing_model",
                        }
                    )
                    continue

                barcode = ""
                if barcode_col is not None and barcode_col - 1 < len(row_values):
                    barcode = self._text(row_values[barcode_col - 1])

                parameters = self._collect_parameters(
                    row_values=row_values,
                    header_names=header_names,
                    fixed_cols={model_col, barcode_col, image_col},
                )

                row_images: list[_RowImage] = []
                if image_col is not None and image_col - 1 < len(row_values):
                    image_cell = self._text(row_values[image_col - 1])
                    for url in self._extract_urls(image_cell):
                        row_images.append(
                            _RowImage(
                                source_type="url",
                                image_url=url,
                                rel_path="",
                                mime_type="",
                                filename="",
                                content=b"",
                            )
                        )
                row_images.extend(embedded_by_row.get(row_idx, []))

                records_by_model[model] = {
                    "sheet": ws.title,
                    "row": row_idx,
                    "barcode": barcode,
                    "parameters": parameters,
                    "images": row_images,
                }

        return {
            "records": records_by_model,
            "errors": errors,
            "rows_scanned": rows_scanned,
        }

    def _extract_embedded_images_by_row(self, ws: Any) -> dict[int, list[_RowImage]]:
        from openpyxl.utils.cell import coordinate_to_tuple

        out: dict[int, list[_RowImage]] = {}
        for image in list(getattr(ws, "_images", []) or []):
            row_idx = None
            anchor = getattr(image, "anchor", None)
            try:
                if anchor is not None and hasattr(anchor, "_from") and getattr(anchor, "_from") is not None:
                    row_idx = int(anchor._from.row) + 1
                elif isinstance(anchor, str):
                    row_idx, _ = coordinate_to_tuple(anchor)
            except Exception:
                row_idx = None
            if row_idx is None or row_idx <= 1:
                continue

            content = self._extract_image_bytes(image)
            if not content:
                continue

            mime_type, ext = self._resolve_image_mime(image)
            filename = self._guess_image_filename(image, ext)
            out.setdefault(row_idx, []).append(
                _RowImage(
                    source_type="embedded",
                    image_url="",
                    rel_path="",
                    mime_type=mime_type,
                    filename=filename,
                    content=content,
                )
            )
        return out

    @staticmethod
    def _extract_image_bytes(image: Any) -> bytes:
        try:
            payload = image._data()
            if isinstance(payload, (bytes, bytearray)):
                return bytes(payload)
        except Exception:
            pass

        ref = getattr(image, "ref", None)
        if isinstance(ref, (bytes, bytearray)):
            return bytes(ref)
        if hasattr(ref, "read"):
            try:
                cursor = ref.tell() if hasattr(ref, "tell") else None
                if hasattr(ref, "seek"):
                    ref.seek(0)
                data = ref.read()
                if cursor is not None and hasattr(ref, "seek"):
                    ref.seek(cursor)
                if isinstance(data, (bytes, bytearray)):
                    return bytes(data)
            except Exception:
                return b""
        return b""

    @staticmethod
    def _resolve_image_mime(image: Any) -> tuple[str, str]:
        img_format = str(getattr(image, "format", "") or "").strip().lower()
        mime_type = _IMAGE_MIME_BY_FORMAT.get(img_format, "image/png")
        ext = _IMAGE_EXT_BY_MIME.get(mime_type, "png")
        return mime_type, ext

    @staticmethod
    def _guess_image_filename(image: Any, ext: str) -> str:
        raw_path = str(getattr(image, "path", "") or "").strip()
        if raw_path:
            base = os.path.basename(raw_path.replace("\\", "/"))
            if "." in base:
                return base
        return f"embedded.{ext}"

    @staticmethod
    def _find_fixed_column(header_names: dict[int, str], aliases: set[str]) -> int | None:
        normalized_aliases = {_normalize_token(x) for x in aliases}
        for col_idx, text in header_names.items():
            if _normalize_token(text) in normalized_aliases:
                return col_idx
        return None

    @staticmethod
    def _iter_header(ws: Any) -> list[Any]:
        max_col = int(ws.max_column or 0)
        if max_col <= 0:
            return []
        return [ws.cell(row=1, column=col_idx).value for col_idx in range(1, max_col + 1)]

    @staticmethod
    def _normalize_header_name(raw: str, index: int) -> str:
        text = str(raw or "").strip()
        if text:
            return text
        return f"参数列{index}"

    @staticmethod
    def _row_is_empty(row_values: list[Any]) -> bool:
        for value in row_values:
            if str(value or "").strip():
                return False
        return True

    @staticmethod
    def _collect_parameters(
        *,
        row_values: list[Any],
        header_names: dict[int, str],
        fixed_cols: set[int | None],
    ) -> dict[str, str]:
        parameters: dict[str, str] = {}
        used_keys: set[str] = set()
        for col_idx, raw in enumerate(row_values, start=1):
            if col_idx in fixed_cols:
                continue
            value = PackageDrawingManager._text(raw)
            if not value:
                continue
            key = str(header_names.get(col_idx) or f"参数列{col_idx}").strip() or f"参数列{col_idx}"
            unique_key = key
            suffix = 2
            while unique_key in used_keys:
                unique_key = f"{key}_{suffix}"
                suffix += 1
            used_keys.add(unique_key)
            parameters[unique_key] = value
        return parameters

    @staticmethod
    def _extract_urls(text: str) -> list[str]:
        raw = str(text or "").strip()
        if not raw:
            return []
        urls = re.findall(r"https?://[^\s,;，；]+", raw, flags=re.IGNORECASE)
        seen: set[str] = set()
        out: list[str] = []
        for item in urls:
            clean = item.strip()
            if clean and clean not in seen:
                seen.add(clean)
                out.append(clean)
        return out

    @staticmethod
    def _safe_model_dir_name(model: str) -> str:
        clean = re.sub(r"[^a-zA-Z0-9._-]+", "_", str(model or "").strip())
        clean = clean.strip("._")
        if not clean:
            clean = "model"
        return clean[:80]

    def _cleanup_old_paths(self, old_paths: list[str]) -> None:
        for rel in old_paths or []:
            rel_clean = str(rel or "").strip().replace("\\", "/")
            if not rel_clean:
                continue
            abs_path = resolve_managed_child_storage_path(
                rel_clean,
                managed_root=self._image_root,
                field_name="package_drawing_images.rel_path",
            )
            try:
                if abs_path.exists():
                    abs_path.unlink()
            except Exception:
                continue

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value).strip()
        return str(value).strip()


def _normalize_token(value: str) -> str:
    return re.sub(r"[\s_-]+", "", str(value or "").strip().lower())
