from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import time

from openpyxl import load_workbook
import xlrd

from backend.app.core.paths import resolve_repo_path

from .models import Company, Department, Employee, OrgStructureRebuildSummary
from .store import OrgDirectoryStore


EXCEL_RELATIVE_PATH = Path("doc") / "上海瑛泰医疗器械股份有限公司在职员工20260403.xls"
HEADER_EMPLOYEE_USER_ID = "员工UserID"
HEADER_EMPLOYEE_NAME = "姓名"
HEADER_EMPLOYEE_EMAIL = "邮箱"
HEADER_EMPLOYEE_NO = "工号"
HEADER_DEPARTMENT_MANAGER = "部门主管"
HEADER_COMPANY = "1级部门"
HEADER_SOURCE_DEPARTMENT_ID = "主部门ID"
DEPARTMENT_LEVEL_HEADERS = tuple(f"{idx}级部门" for idx in range(2, 8))


@dataclass(frozen=True)
class _ParsedCompany:
    name: str
    source_key: str


@dataclass(frozen=True)
class _ParsedDepartment:
    name: str
    company_source_key: str
    parent_source_key: str | None
    source_key: str
    source_department_id: str | None
    level_no: int
    path_name: str
    sort_order: int


@dataclass(frozen=True)
class _ParsedEmployee:
    employee_user_id: str
    name: str
    email: str | None
    employee_no: str | None
    department_manager_name: str | None
    is_department_manager: bool
    company_source_key: str
    department_source_key: str | None
    source_key: str
    path_name: str
    sort_order: int


class OrgStructureManager:
    def __init__(self, *, store: OrgDirectoryStore, excel_path: str | Path | None = None):
        if store is None:
            raise RuntimeError("org_directory_store_unavailable")
        self._store = store
        self._excel_path = resolve_repo_path(excel_path or EXCEL_RELATIVE_PATH)

    @property
    def excel_path(self) -> Path:
        return self._excel_path

    def list_companies(self) -> list[Company]:
        return self._store.list_companies()

    def list_departments_flat(self) -> list[Department]:
        return self._store.list_departments()

    def list_employees(self) -> list[Employee]:
        return self._store.list_employees()

    def get_company(self, company_id: int) -> Company | None:
        return self._store.get_company(company_id)

    def get_department(self, department_id: int) -> Department | None:
        return self._store.get_department(department_id)

    def get_employee(self, employee_id: int) -> Employee | None:
        return self._store.get_employee(employee_id)

    def list_audit_logs(
        self,
        *,
        entity_type: str | None = None,
        action: str | None = None,
        limit: int = 200,
    ):
        return self._store.list_audit_logs(entity_type=entity_type, action=action, limit=limit)

    def get_tree(self) -> list[dict]:
        companies = self.list_companies()
        departments = self.list_departments_flat()
        employees = self.list_employees()

        company_by_id = {item.company_id: item for item in companies}
        department_by_id = {item.department_id: item for item in departments}
        children_by_parent: dict[int, list[dict]] = {}
        roots_by_company: dict[int, list[dict]] = {}
        people_by_department: dict[int, list[dict]] = {}
        people_by_company: dict[int, list[dict]] = {}

        for department in departments:
            node = self._department_node(department)
            if department.parent_department_id is None:
                roots_by_company.setdefault(int(department.company_id or 0), []).append(node)
            else:
                children_by_parent.setdefault(department.parent_department_id, []).append(node)

        for employee in employees:
            company = company_by_id.get(int(employee.company_id or 0))
            if company is None:
                raise RuntimeError(f"org_employee_company_missing:{employee.employee_user_id}")
            department = None
            if employee.department_id is not None:
                department = department_by_id.get(employee.department_id)
                if department is None:
                    raise RuntimeError(f"org_employee_department_missing:{employee.employee_user_id}")
            node = self._employee_node(employee, company=company, department=department)
            if employee.department_id is None:
                people_by_company.setdefault(company.company_id, []).append(node)
            else:
                people_by_department.setdefault(employee.department_id, []).append(node)

        def attach_children(nodes: list[dict]) -> list[dict]:
            attached: list[dict] = []
            for node in nodes:
                child_nodes = children_by_parent.get(node["id"], [])
                node["children"] = attach_children(child_nodes) + people_by_department.get(node["id"], [])
                attached.append(node)
            return attached

        tree: list[dict] = []
        for company in companies:
            root_children = roots_by_company.get(company.company_id, [])
            tree.append(
                {
                    "id": company.company_id,
                    "node_type": "company",
                    "name": company.name,
                    "path_name": company.name,
                    "source_key": company.source_key,
                    "company_id": company.company_id,
                    "department_id": None,
                    "parent_department_id": None,
                    "level_no": 1,
                    "source_department_id": None,
                    "employee_user_id": None,
                    "email": None,
                    "employee_no": None,
                    "department_manager_name": None,
                    "is_department_manager": False,
                    "created_at_ms": company.created_at_ms,
                    "updated_at_ms": company.updated_at_ms,
                    "children": attach_children(root_children) + people_by_company.get(company.company_id, []),
                }
            )
        return tree

    def rebuild_from_excel(
        self,
        *,
        actor_user_id: str,
        excel_path: str | Path | None = None,
        source_label: str | None = None,
    ) -> OrgStructureRebuildSummary:
        actor = str(actor_user_id or "").strip()
        if not actor:
            raise RuntimeError("actor_user_id_required")
        resolved_excel_path = self._resolve_excel_path(excel_path)
        source_display = str(source_label or resolved_excel_path)
        companies, departments, employees = self._parse_excel(excel_path=resolved_excel_path)

        company_keys = {item.source_key for item in companies}
        department_keys = {item.source_key for item in departments}
        employee_keys = {item.source_key for item in employees}
        completed_at_ms = int(time.time() * 1000)

        company_created = 0
        company_updated = 0
        company_deleted = 0
        department_created = 0
        department_updated = 0
        department_deleted = 0
        employee_created = 0
        employee_updated = 0
        employee_deleted = 0
        users_company_cleared = 0
        users_department_cleared = 0

        conn = self._store._get_connection()
        try:
            conn.execute("BEGIN IMMEDIATE")

            existing_companies = {
                row["source_key"]: Company(
                    company_id=int(row["company_id"]),
                    name=str(row["name"] or ""),
                    source_key=str(row["source_key"] or ""),
                    created_at_ms=int(row["created_at_ms"]),
                    updated_at_ms=int(row["updated_at_ms"]),
                )
                for row in conn.execute(
                    """
                    SELECT company_id, name, source_key, created_at_ms, updated_at_ms
                    FROM companies
                    WHERE source_key IS NOT NULL
                    """
                ).fetchall()
            }

            company_id_by_source_key: dict[str, int] = {}
            for item in companies:
                current = existing_companies.get(item.source_key)
                if current is None:
                    cur = conn.execute(
                        """
                        INSERT INTO companies (name, source_key, created_at_ms, updated_at_ms)
                        VALUES (?, ?, ?, ?)
                        """,
                        (item.name, item.source_key, completed_at_ms, completed_at_ms),
                    )
                    company_id = int(cur.lastrowid)
                    company_created += 1
                    self._store._log(
                        conn,
                        entity_type="company",
                        action="create",
                        entity_id=company_id,
                        before_name=None,
                        after_name=item.name,
                        actor_user_id=actor,
                    )
                else:
                    company_id = current.company_id
                    if current.name != item.name:
                        conn.execute(
                            """
                            UPDATE companies
                            SET name = ?, source_key = ?, updated_at_ms = ?
                            WHERE company_id = ?
                            """,
                            (item.name, item.source_key, completed_at_ms, company_id),
                        )
                        company_updated += 1
                        self._store._log(
                            conn,
                            entity_type="company",
                            action="update",
                            entity_id=company_id,
                            before_name=current.name,
                            after_name=item.name,
                            actor_user_id=actor,
                        )
                    elif current.source_key != item.source_key:
                        conn.execute(
                            "UPDATE companies SET source_key = ?, updated_at_ms = ? WHERE company_id = ?",
                            (item.source_key, completed_at_ms, company_id),
                        )
                company_id_by_source_key[item.source_key] = company_id

            existing_departments = {
                row["source_key"]: Department(
                    department_id=int(row["department_id"]),
                    name=str(row["name"] or ""),
                    company_id=(int(row["company_id"]) if row["company_id"] is not None else None),
                    parent_department_id=(
                        int(row["parent_department_id"]) if row["parent_department_id"] is not None else None
                    ),
                    source_key=str(row["source_key"] or ""),
                    source_department_id=(
                        str(row["source_department_id"]).strip() if row["source_department_id"] is not None else None
                    ),
                    level_no=int(row["level_no"] or 0),
                    path_name=str(row["path_name"] or ""),
                    sort_order=int(row["sort_order"] or 0),
                    created_at_ms=int(row["created_at_ms"]),
                    updated_at_ms=int(row["updated_at_ms"]),
                )
                for row in conn.execute(
                    """
                    SELECT
                        department_id,
                        name,
                        company_id,
                        parent_department_id,
                        source_key,
                        source_department_id,
                        level_no,
                        path_name,
                        sort_order,
                        created_at_ms,
                        updated_at_ms
                    FROM departments
                    """
                ).fetchall()
            }

            department_id_by_source_key: dict[str, int] = {}
            for item in departments:
                company_id = company_id_by_source_key[item.company_source_key]
                parent_department_id = (
                    department_id_by_source_key[item.parent_source_key] if item.parent_source_key is not None else None
                )
                current = existing_departments.get(item.source_key)
                if current is None:
                    cur = conn.execute(
                        """
                        INSERT INTO departments (
                            name,
                            company_id,
                            parent_department_id,
                            source_key,
                            source_department_id,
                            level_no,
                            path_name,
                            sort_order,
                            created_at_ms,
                            updated_at_ms
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item.name,
                            company_id,
                            parent_department_id,
                            item.source_key,
                            item.source_department_id,
                            item.level_no,
                            item.path_name,
                            item.sort_order,
                            completed_at_ms,
                            completed_at_ms,
                        ),
                    )
                    department_id = int(cur.lastrowid)
                    department_created += 1
                    self._store._log(
                        conn,
                        entity_type="department",
                        action="create",
                        entity_id=department_id,
                        before_name=None,
                        after_name=item.path_name,
                        actor_user_id=actor,
                    )
                else:
                    department_id = current.department_id
                    needs_update = (
                        current.name != item.name
                        or current.company_id != company_id
                        or current.parent_department_id != parent_department_id
                        or current.source_department_id != item.source_department_id
                        or current.level_no != item.level_no
                        or current.path_name != item.path_name
                        or current.sort_order != item.sort_order
                    )
                    if needs_update:
                        conn.execute(
                            """
                            UPDATE departments
                            SET
                                name = ?,
                                company_id = ?,
                                parent_department_id = ?,
                                source_key = ?,
                                source_department_id = ?,
                                level_no = ?,
                                path_name = ?,
                                sort_order = ?,
                                updated_at_ms = ?
                            WHERE department_id = ?
                            """,
                            (
                                item.name,
                                company_id,
                                parent_department_id,
                                item.source_key,
                                item.source_department_id,
                                item.level_no,
                                item.path_name,
                                item.sort_order,
                                completed_at_ms,
                                department_id,
                            ),
                        )
                        department_updated += 1
                        self._store._log(
                            conn,
                            entity_type="department",
                            action="update",
                            entity_id=department_id,
                            before_name=current.path_name or current.name,
                            after_name=item.path_name,
                            actor_user_id=actor,
                        )
                department_id_by_source_key[item.source_key] = department_id

            existing_employees = {
                row["source_key"]: Employee(
                    employee_id=int(row["employee_id"]),
                    employee_user_id=str(row["employee_user_id"] or ""),
                    name=str(row["name"] or ""),
                    email=(str(row["email"]).strip() if row["email"] is not None and str(row["email"]).strip() else None),
                    employee_no=(
                        str(row["employee_no"]).strip()
                        if row["employee_no"] is not None and str(row["employee_no"]).strip()
                        else None
                    ),
                    department_manager_name=(
                        str(row["department_manager_name"]).strip()
                        if row["department_manager_name"] is not None and str(row["department_manager_name"]).strip()
                        else None
                    ),
                    is_department_manager=bool(int(row["is_department_manager"] or 0)),
                    company_id=(int(row["company_id"]) if row["company_id"] is not None else None),
                    department_id=(int(row["department_id"]) if row["department_id"] is not None else None),
                    source_key=str(row["source_key"] or ""),
                    sort_order=int(row["sort_order"] or 0),
                    created_at_ms=int(row["created_at_ms"]),
                    updated_at_ms=int(row["updated_at_ms"]),
                )
                for row in conn.execute(
                    """
                    SELECT
                        employee_id,
                        employee_user_id,
                        name,
                        email,
                        employee_no,
                        department_manager_name,
                        is_department_manager,
                        company_id,
                        department_id,
                        source_key,
                        sort_order,
                        created_at_ms,
                        updated_at_ms
                    FROM org_employees
                    """
                ).fetchall()
            }

            for item in employees:
                company_id = company_id_by_source_key[item.company_source_key]
                department_id = (
                    department_id_by_source_key[item.department_source_key]
                    if item.department_source_key is not None
                    else None
                )
                current = existing_employees.get(item.source_key)
                if current is None:
                    conn.execute(
                        """
                        INSERT INTO org_employees (
                            employee_user_id,
                            name,
                            email,
                            employee_no,
                            department_manager_name,
                            is_department_manager,
                            company_id,
                            department_id,
                            source_key,
                            sort_order,
                            created_at_ms,
                            updated_at_ms
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item.employee_user_id,
                            item.name,
                            item.email,
                            item.employee_no,
                            item.department_manager_name,
                            1 if item.is_department_manager else 0,
                            company_id,
                            department_id,
                            item.source_key,
                            item.sort_order,
                            completed_at_ms,
                            completed_at_ms,
                        ),
                    )
                    employee_created += 1
                else:
                    needs_update = (
                        current.employee_user_id != item.employee_user_id
                        or current.name != item.name
                        or current.email != item.email
                        or current.employee_no != item.employee_no
                        or current.department_manager_name != item.department_manager_name
                        or current.is_department_manager != item.is_department_manager
                        or current.company_id != company_id
                        or current.department_id != department_id
                        or current.sort_order != item.sort_order
                    )
                    if needs_update:
                        conn.execute(
                            """
                            UPDATE org_employees
                            SET
                                employee_user_id = ?,
                                name = ?,
                                email = ?,
                                employee_no = ?,
                                department_manager_name = ?,
                                is_department_manager = ?,
                                company_id = ?,
                                department_id = ?,
                                source_key = ?,
                                sort_order = ?,
                                updated_at_ms = ?
                            WHERE employee_id = ?
                            """,
                            (
                                item.employee_user_id,
                                item.name,
                                item.email,
                                item.employee_no,
                                item.department_manager_name,
                                1 if item.is_department_manager else 0,
                                company_id,
                                department_id,
                                item.source_key,
                                item.sort_order,
                                completed_at_ms,
                                current.employee_id,
                            ),
                        )
                        employee_updated += 1

            stale_employee_rows = [
                row
                for row in conn.execute(
                    """
                    SELECT employee_id, source_key
                    FROM org_employees
                    """
                ).fetchall()
                if str(row["source_key"] or "") not in employee_keys
            ]
            stale_employee_ids = [int(row["employee_id"]) for row in stale_employee_rows]
            if stale_employee_ids:
                for batch in self._iter_chunks(stale_employee_ids, 500):
                    conn.execute(
                        "DELETE FROM org_employees WHERE employee_id IN ({placeholders})".format(
                            placeholders=self._placeholders(len(batch))
                        ),
                        tuple(batch),
                    )
                employee_deleted = len(stale_employee_ids)

            stale_departments = conn.execute(
                """
                SELECT
                    department_id,
                    name,
                    path_name
                FROM departments
                WHERE source_key NOT IN ({placeholders})
                ORDER BY level_no DESC, department_id DESC
                """.format(placeholders=self._placeholders(len(department_keys))),
                tuple(sorted(department_keys)),
            ).fetchall() if department_keys else conn.execute(
                """
                SELECT department_id, name, path_name
                FROM departments
                ORDER BY level_no DESC, department_id DESC
                """
            ).fetchall()

            stale_companies = conn.execute(
                """
                SELECT company_id, name
                FROM companies
                WHERE source_key NOT IN ({placeholders})
                ORDER BY company_id DESC
                """.format(placeholders=self._placeholders(len(company_keys))),
                tuple(sorted(company_keys)),
            ).fetchall() if company_keys else conn.execute(
                """
                SELECT company_id, name
                FROM companies
                ORDER BY company_id DESC
                """
            ).fetchall()

            stale_department_ids = [int(row["department_id"]) for row in stale_departments]
            if stale_department_ids:
                users_department_cleared = conn.execute(
                    """
                    UPDATE users
                    SET department_id = NULL
                    WHERE department_id IN ({placeholders})
                    """.format(placeholders=self._placeholders(len(stale_department_ids))),
                    tuple(stale_department_ids),
                ).rowcount
                for row in stale_departments:
                    self._store._log(
                        conn,
                        entity_type="department",
                        action="delete",
                        entity_id=int(row["department_id"]),
                        before_name=str(row["path_name"] or row["name"] or ""),
                        after_name=None,
                        actor_user_id=actor,
                    )
                conn.execute(
                    "DELETE FROM departments WHERE department_id IN ({placeholders})".format(
                        placeholders=self._placeholders(len(stale_department_ids))
                    ),
                    tuple(stale_department_ids),
                )
                department_deleted = len(stale_department_ids)

            stale_company_ids = [int(row["company_id"]) for row in stale_companies]
            if stale_company_ids:
                users_company_cleared = conn.execute(
                    """
                    UPDATE users
                    SET company_id = NULL
                    WHERE company_id IN ({placeholders})
                    """.format(placeholders=self._placeholders(len(stale_company_ids))),
                    tuple(stale_company_ids),
                ).rowcount
                for row in stale_companies:
                    self._store._log(
                        conn,
                        entity_type="company",
                        action="delete",
                        entity_id=int(row["company_id"]),
                        before_name=str(row["name"] or ""),
                        after_name=None,
                        actor_user_id=actor,
                    )
                conn.execute(
                    "DELETE FROM companies WHERE company_id IN ({placeholders})".format(
                        placeholders=self._placeholders(len(stale_company_ids))
                    ),
                    tuple(stale_company_ids),
                )
                company_deleted = len(stale_company_ids)

            summary_text = (
                f"companies={len(companies)}, departments={len(departments)}, employees={len(employees)}, "
                f"company_changes=+{company_created}/~{company_updated}/-{company_deleted}, "
                f"department_changes=+{department_created}/~{department_updated}/-{department_deleted}, "
                f"employee_changes=+{employee_created}/~{employee_updated}/-{employee_deleted}, "
                f"user_refs_cleared=company:{users_company_cleared},department:{users_department_cleared}"
            )
            self._store._log(
                conn,
                entity_type="org_structure",
                action="rebuild",
                entity_id=None,
                before_name=source_display,
                after_name=summary_text,
                actor_user_id=actor,
            )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

        return OrgStructureRebuildSummary(
            excel_path=source_display,
            company_count=len(companies),
            department_count=len(departments),
            employee_count=len(employees),
            companies_created=company_created,
            companies_updated=company_updated,
            companies_deleted=company_deleted,
            departments_created=department_created,
            departments_updated=department_updated,
            departments_deleted=department_deleted,
            employees_created=employee_created,
            employees_updated=employee_updated,
            employees_deleted=employee_deleted,
            users_company_cleared=users_company_cleared,
            users_department_cleared=users_department_cleared,
            completed_at_ms=completed_at_ms,
        )

    def _department_node(self, department: Department) -> dict:
        return {
            "id": department.department_id,
            "node_type": "department",
            "name": department.name,
            "path_name": department.path_name,
            "source_key": department.source_key,
            "company_id": department.company_id,
            "department_id": department.department_id,
            "parent_department_id": department.parent_department_id,
            "level_no": department.level_no,
            "source_department_id": department.source_department_id,
            "employee_user_id": None,
            "email": None,
            "employee_no": None,
            "department_manager_name": None,
            "is_department_manager": False,
            "created_at_ms": department.created_at_ms,
            "updated_at_ms": department.updated_at_ms,
            "children": [],
        }

    def _employee_node(self, employee: Employee, *, company: Company, department: Department | None) -> dict:
        parent_path = department.path_name if department is not None else company.name
        return {
            "id": employee.employee_id,
            "node_type": "person",
            "name": employee.name,
            "path_name": f"{parent_path} / {employee.name}",
            "source_key": employee.source_key,
            "company_id": employee.company_id,
            "department_id": employee.department_id,
            "parent_department_id": employee.department_id,
            "level_no": (department.level_no + 1) if department is not None else 2,
            "source_department_id": None,
            "employee_user_id": employee.employee_user_id,
            "email": employee.email,
            "employee_no": employee.employee_no,
            "department_manager_name": employee.department_manager_name,
            "is_department_manager": employee.is_department_manager,
            "created_at_ms": employee.created_at_ms,
            "updated_at_ms": employee.updated_at_ms,
            "children": [],
        }

    def _parse_excel(
        self,
        *,
        excel_path: str | Path | None = None,
    ) -> tuple[list[_ParsedCompany], list[_ParsedDepartment], list[_ParsedEmployee]]:
        rows = self._load_excel_rows(excel_path=excel_path)
        if not rows:
            raise RuntimeError("org_structure_excel_empty")

        headers = [self._normalize_cell_value(value) for value in rows[0]]
        header_index = {header: idx for idx, header in enumerate(headers) if header}

        required_headers = {
            HEADER_EMPLOYEE_USER_ID,
            HEADER_EMPLOYEE_NAME,
            HEADER_EMPLOYEE_EMAIL,
            HEADER_EMPLOYEE_NO,
            HEADER_DEPARTMENT_MANAGER,
            HEADER_COMPANY,
            HEADER_SOURCE_DEPARTMENT_ID,
            *DEPARTMENT_LEVEL_HEADERS,
        }
        missing_headers = sorted(header for header in required_headers if header not in header_index)
        if missing_headers:
            raise RuntimeError(f"org_structure_excel_headers_missing:{','.join(missing_headers)}")

        company_order: list[str] = []
        seen_company_keys: set[str] = set()
        department_order: list[_ParsedDepartment] = []
        seen_department_keys: set[str] = set()
        employee_order: list[_ParsedEmployee] = []
        seen_employee_keys: set[str] = set()
        sibling_order: dict[tuple[str, str | None], int] = {}
        employee_sibling_order: dict[tuple[str, str | None], int] = {}
        path_to_source_department_id: dict[str, str] = {}
        source_department_id_to_path: dict[str, str] = {}

        for row_idx, row_values in enumerate(rows[1:], start=1):
            employee_user_id = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_EMPLOYEE_USER_ID])
            )
            employee_name = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_EMPLOYEE_NAME])
            )
            employee_email = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_EMPLOYEE_EMAIL])
            )
            employee_no = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_EMPLOYEE_NO])
            )
            department_manager_name = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_DEPARTMENT_MANAGER])
            )
            company_name = self._normalize_cell_value(self._row_value(row_values, header_index[HEADER_COMPANY]))
            level_values = [
                self._normalize_cell_value(self._row_value(row_values, header_index[level_header]))
                for level_header in DEPARTMENT_LEVEL_HEADERS
            ]
            source_department_id = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_SOURCE_DEPARTMENT_ID])
            )
            non_empty_levels = [value for value in level_values if value]
            first_blank_seen = False
            for level_name in level_values:
                if not level_name:
                    first_blank_seen = True
                    continue
                if first_blank_seen:
                    raise RuntimeError(f"org_structure_excel_level_gap:row_{row_idx + 1}")

            if not any([employee_user_id, employee_name, employee_email, employee_no, company_name, *non_empty_levels]):
                continue
            if not company_name:
                raise RuntimeError(f"org_structure_excel_company_required:row_{row_idx + 1}")
            if not employee_user_id:
                raise RuntimeError(f"org_structure_excel_employee_user_id_required:row_{row_idx + 1}")
            if not employee_name:
                raise RuntimeError(f"org_structure_excel_employee_name_required:row_{row_idx + 1}")

            if company_name not in seen_company_keys:
                seen_company_keys.add(company_name)
                company_order.append(company_name)

            parent_source_key: str | None = None
            path_parts = [company_name]
            for level_offset, level_name in enumerate(level_values, start=2):
                if not level_name:
                    break
                path_parts.append(level_name)
                source_key = "/".join(path_parts)
                if source_key in seen_department_keys:
                    parent_source_key = source_key
                    continue

                sort_key = (company_name, parent_source_key)
                sort_order = sibling_order.get(sort_key, 0)
                sibling_order[sort_key] = sort_order + 1

                department_order.append(
                    _ParsedDepartment(
                        name=level_name,
                        company_source_key=company_name,
                        parent_source_key=parent_source_key,
                        source_key=source_key,
                        source_department_id=None,
                        level_no=level_offset,
                        path_name=" / ".join(path_parts),
                        sort_order=sort_order,
                    )
                )
                seen_department_keys.add(source_key)
                parent_source_key = source_key

            terminal_source_key: str | None = None
            if non_empty_levels:
                terminal_source_key = "/".join([company_name, *non_empty_levels])
                if not source_department_id:
                    raise RuntimeError(f"org_structure_excel_source_department_id_required:row_{row_idx + 1}")
                previous_path = source_department_id_to_path.get(source_department_id)
                if previous_path is not None and previous_path != terminal_source_key:
                    raise RuntimeError(f"org_structure_excel_source_department_id_conflict:{source_department_id}")
                previous_source_department_id = path_to_source_department_id.get(terminal_source_key)
                if previous_source_department_id is not None and previous_source_department_id != source_department_id:
                    raise RuntimeError(f"org_structure_excel_path_conflict:{terminal_source_key}")
                source_department_id_to_path[source_department_id] = terminal_source_key
                path_to_source_department_id[terminal_source_key] = source_department_id

            employee_source_key = employee_user_id
            if employee_source_key in seen_employee_keys:
                raise RuntimeError(f"org_structure_excel_employee_user_id_conflict:{employee_user_id}")
            seen_employee_keys.add(employee_source_key)

            employee_sort_key = (company_name, terminal_source_key)
            employee_sort_order = employee_sibling_order.get(employee_sort_key, 0)
            employee_sibling_order[employee_sort_key] = employee_sort_order + 1
            employee_order.append(
                _ParsedEmployee(
                    employee_user_id=employee_user_id,
                    name=employee_name,
                    email=employee_email or None,
                    employee_no=employee_no or None,
                    department_manager_name=department_manager_name or None,
                    is_department_manager=bool(department_manager_name and department_manager_name == employee_name),
                    company_source_key=company_name,
                    department_source_key=terminal_source_key,
                    source_key=employee_source_key,
                    path_name=" / ".join([company_name, *non_empty_levels, employee_name]),
                    sort_order=employee_sort_order,
                )
            )

        department_items: list[_ParsedDepartment] = []
        for item in department_order:
            terminal_source_department_id = path_to_source_department_id.get(item.source_key)
            department_items.append(
                _ParsedDepartment(
                    name=item.name,
                    company_source_key=item.company_source_key,
                    parent_source_key=item.parent_source_key,
                    source_key=item.source_key,
                    source_department_id=terminal_source_department_id,
                    level_no=item.level_no,
                    path_name=item.path_name,
                    sort_order=item.sort_order,
                )
            )

        company_items = [_ParsedCompany(name=name, source_key=name) for name in company_order]
        return company_items, department_items, employee_order

    def _load_excel_rows(self, *, excel_path: str | Path | None = None) -> list[list[object]]:
        resolved_excel_path = self._resolve_excel_path(excel_path)
        if not resolved_excel_path.exists():
            raise RuntimeError(f"org_structure_excel_not_found:{resolved_excel_path}")

        suffix = resolved_excel_path.suffix.lower()
        if suffix == ".xls":
            workbook = xlrd.open_workbook(filename=str(resolved_excel_path))
            try:
                sheet = workbook.sheet_by_index(0)
            except IndexError as exc:
                raise RuntimeError("org_structure_excel_sheet_missing") from exc
            return [
                [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                for row_idx in range(sheet.nrows)
            ]

        if suffix == ".xlsx":
            workbook = load_workbook(filename=str(resolved_excel_path), read_only=True, data_only=True)
            try:
                sheet = workbook.worksheets[0]
            except IndexError as exc:
                workbook.close()
                raise RuntimeError("org_structure_excel_sheet_missing") from exc
            try:
                return [list(row) for row in sheet.iter_rows(values_only=True)]
            finally:
                workbook.close()

        raise RuntimeError(f"org_structure_excel_extension_invalid:{suffix or resolved_excel_path.name}")

    def _resolve_excel_path(self, excel_path: str | Path | None) -> Path:
        if excel_path is None:
            return self._excel_path
        candidate = Path(excel_path)
        if candidate.is_absolute():
            return candidate
        return resolve_repo_path(candidate)

    @staticmethod
    def _row_value(row_values: list[object], index: int):
        if index < 0 or index >= len(row_values):
            return ""
        return row_values[index]

    @staticmethod
    def _normalize_cell_value(value) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        if text.endswith(".0"):
            integer_text = text[:-2]
            if integer_text.isdigit():
                return integer_text
        return text

    @staticmethod
    def _placeholders(size: int) -> str:
        return ",".join("?" for _ in range(size))

    @staticmethod
    def _iter_chunks(values: list[int], size: int):
        for idx in range(0, len(values), size):
            yield values[idx : idx + size]
