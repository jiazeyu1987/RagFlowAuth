from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
import xlrd

from backend.app.core.paths import resolve_repo_path

from .rebuild_types import ParsedCompany, ParsedDepartment, ParsedEmployee, ParsedOrgStructure


EXCEL_RELATIVE_PATH = Path("doc") / "上海瑛泰医疗器械股份有限公司在职员工20260403.xls"
HEADER_EMPLOYEE_USER_ID = "员工UserID"
HEADER_EMPLOYEE_NAME = "姓名"
HEADER_EMPLOYEE_EMAIL = "邮箱"
HEADER_EMPLOYEE_NO = "工号"
HEADER_DEPARTMENT_MANAGER = "部门主管"
HEADER_COMPANY = "1级部门"
HEADER_SOURCE_DEPARTMENT_ID = "主部门ID"
DEPARTMENT_LEVEL_HEADERS = tuple(f"{idx}级部门" for idx in range(2, 8))


class OrgStructureExcelParser:
    def __init__(self, *, excel_path: str | Path | None = None):
        self._excel_path = resolve_repo_path(excel_path or EXCEL_RELATIVE_PATH)

    @property
    def excel_path(self) -> Path:
        return self._excel_path

    def resolve_excel_path(self, excel_path: str | Path | None) -> Path:
        if excel_path is None:
            return self._excel_path
        candidate = Path(excel_path)
        if candidate.is_absolute():
            return candidate
        return resolve_repo_path(candidate)

    def parse(self, *, excel_path: str | Path | None = None) -> ParsedOrgStructure:
        resolved_excel_path = self.resolve_excel_path(excel_path)
        rows = self._load_excel_rows(excel_path=resolved_excel_path)
        if not rows:
            raise RuntimeError("org_structure_excel_empty")

        headers = [self._normalize_cell_value(value) for value in rows[0]]
        header_index = {header: idx for idx, header in enumerate(headers) if header}

        required_headers = {
            HEADER_EMPLOYEE_USER_ID,
            HEADER_EMPLOYEE_NAME,
            HEADER_EMPLOYEE_EMAIL,
            HEADER_EMPLOYEE_NO,
            HEADER_DEPARTMENT_MANAGER,
            HEADER_COMPANY,
            HEADER_SOURCE_DEPARTMENT_ID,
            *DEPARTMENT_LEVEL_HEADERS,
        }
        missing_headers = sorted(header for header in required_headers if header not in header_index)
        if missing_headers:
            raise RuntimeError(f"org_structure_excel_headers_missing:{','.join(missing_headers)}")

        company_order: list[str] = []
        seen_company_keys: set[str] = set()
        department_order: list[ParsedDepartment] = []
        seen_department_keys: set[str] = set()
        employee_order: list[ParsedEmployee] = []
        seen_employee_keys: set[str] = set()
        sibling_order: dict[tuple[str, str | None], int] = {}
        employee_sibling_order: dict[tuple[str, str | None], int] = {}
        path_to_source_department_id: dict[str, str] = {}
        source_department_id_to_path: dict[str, str] = {}

        for row_idx, row_values in enumerate(rows[1:], start=1):
            employee_user_id = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_EMPLOYEE_USER_ID])
            )
            employee_name = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_EMPLOYEE_NAME])
            )
            employee_email = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_EMPLOYEE_EMAIL])
            )
            employee_no = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_EMPLOYEE_NO])
            )
            department_manager_name = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_DEPARTMENT_MANAGER])
            )
            company_name = self._normalize_cell_value(self._row_value(row_values, header_index[HEADER_COMPANY]))
            level_values = [
                self._normalize_cell_value(self._row_value(row_values, header_index[level_header]))
                for level_header in DEPARTMENT_LEVEL_HEADERS
            ]
            source_department_id = self._normalize_cell_value(
                self._row_value(row_values, header_index[HEADER_SOURCE_DEPARTMENT_ID])
            )
            non_empty_levels = [value for value in level_values if value]
            first_blank_seen = False
            for level_name in level_values:
                if not level_name:
                    first_blank_seen = True
                    continue
                if first_blank_seen:
                    raise RuntimeError(f"org_structure_excel_level_gap:row_{row_idx + 1}")

            if not any([employee_user_id, employee_name, employee_email, employee_no, company_name, *non_empty_levels]):
                continue
            if not company_name:
                raise RuntimeError(f"org_structure_excel_company_required:row_{row_idx + 1}")
            if not employee_user_id:
                raise RuntimeError(f"org_structure_excel_employee_user_id_required:row_{row_idx + 1}")
            if not employee_name:
                raise RuntimeError(f"org_structure_excel_employee_name_required:row_{row_idx + 1}")

            if company_name not in seen_company_keys:
                seen_company_keys.add(company_name)
                company_order.append(company_name)

            parent_source_key: str | None = None
            path_parts = [company_name]
            for level_offset, level_name in enumerate(level_values, start=2):
                if not level_name:
                    break
                path_parts.append(level_name)
                source_key = "/".join(path_parts)
                if source_key in seen_department_keys:
                    parent_source_key = source_key
                    continue

                sort_key = (company_name, parent_source_key)
                sort_order = sibling_order.get(sort_key, 0)
                sibling_order[sort_key] = sort_order + 1

                department_order.append(
                    ParsedDepartment(
                        name=level_name,
                        company_source_key=company_name,
                        parent_source_key=parent_source_key,
                        source_key=source_key,
                        source_department_id=None,
                        level_no=level_offset,
                        path_name=" / ".join(path_parts),
                        sort_order=sort_order,
                    )
                )
                seen_department_keys.add(source_key)
                parent_source_key = source_key

            terminal_source_key: str | None = None
            if non_empty_levels:
                terminal_source_key = "/".join([company_name, *non_empty_levels])
                if not source_department_id:
                    raise RuntimeError(f"org_structure_excel_source_department_id_required:row_{row_idx + 1}")
                previous_path = source_department_id_to_path.get(source_department_id)
                if previous_path is not None and previous_path != terminal_source_key:
                    raise RuntimeError(f"org_structure_excel_source_department_id_conflict:{source_department_id}")
                previous_source_department_id = path_to_source_department_id.get(terminal_source_key)
                if previous_source_department_id is not None and previous_source_department_id != source_department_id:
                    raise RuntimeError(f"org_structure_excel_path_conflict:{terminal_source_key}")
                source_department_id_to_path[source_department_id] = terminal_source_key
                path_to_source_department_id[terminal_source_key] = source_department_id

            employee_source_key = employee_user_id
            if employee_source_key in seen_employee_keys:
                raise RuntimeError(f"org_structure_excel_employee_user_id_conflict:{employee_user_id}")
            seen_employee_keys.add(employee_source_key)

            employee_sort_key = (company_name, terminal_source_key)
            employee_sort_order = employee_sibling_order.get(employee_sort_key, 0)
            employee_sibling_order[employee_sort_key] = employee_sort_order + 1
            employee_order.append(
                ParsedEmployee(
                    employee_user_id=employee_user_id,
                    name=employee_name,
                    email=employee_email or None,
                    employee_no=employee_no or None,
                    department_manager_name=department_manager_name or None,
                    is_department_manager=bool(department_manager_name and department_manager_name == employee_name),
                    company_source_key=company_name,
                    department_source_key=terminal_source_key,
                    source_key=employee_source_key,
                    path_name=" / ".join([company_name, *non_empty_levels, employee_name]),
                    sort_order=employee_sort_order,
                )
            )

        department_items: list[ParsedDepartment] = []
        for item in department_order:
            terminal_source_department_id = path_to_source_department_id.get(item.source_key)
            department_items.append(
                ParsedDepartment(
                    name=item.name,
                    company_source_key=item.company_source_key,
                    parent_source_key=item.parent_source_key,
                    source_key=item.source_key,
                    source_department_id=terminal_source_department_id,
                    level_no=item.level_no,
                    path_name=item.path_name,
                    sort_order=item.sort_order,
                )
            )

        company_items = [ParsedCompany(name=name, source_key=name) for name in company_order]
        return ParsedOrgStructure(
            companies=company_items,
            departments=department_items,
            employees=employee_order,
        )

    def _load_excel_rows(self, *, excel_path: str | Path | None = None) -> list[list[object]]:
        resolved_excel_path = self.resolve_excel_path(excel_path)
        if not resolved_excel_path.exists():
            raise RuntimeError(f"org_structure_excel_not_found:{resolved_excel_path}")

        suffix = resolved_excel_path.suffix.lower()
        if suffix == ".xls":
            workbook = xlrd.open_workbook(filename=str(resolved_excel_path))
            try:
                sheet = workbook.sheet_by_index(0)
            except IndexError as exc:
                raise RuntimeError("org_structure_excel_sheet_missing") from exc
            return [
                [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                for row_idx in range(sheet.nrows)
            ]

        if suffix == ".xlsx":
            workbook = load_workbook(filename=str(resolved_excel_path), read_only=True, data_only=True)
            try:
                sheet = workbook.worksheets[0]
            except IndexError as exc:
                workbook.close()
                raise RuntimeError("org_structure_excel_sheet_missing") from exc
            try:
                return [list(row) for row in sheet.iter_rows(values_only=True)]
            finally:
                workbook.close()

        raise RuntimeError(f"org_structure_excel_extension_invalid:{suffix or resolved_excel_path.name}")

    @staticmethod
    def _row_value(row_values: list[object], index: int):
        if index < 0 or index >= len(row_values):
            return ""
        return row_values[index]

    @staticmethod
    def _normalize_cell_value(value) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        if text.endswith(".0"):
            integer_text = text[:-2]
            if integer_text.isdigit():
                return integer_text
        return text
